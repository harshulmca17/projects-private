import cognitive_face as CF
personGroupId = 'test1'
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

import time
os.chdir("C:/Users/harsh/Desktop/DBA/project")

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = 'reports%s.xlsx' % currentDate)
sheet = wb.get_sheet_by_name('Cse15')
print("1")
#print(type(currentDate))
def getDateColumn():
	for i in range(1, sheet.max_column+1):
		col = get_column_letter(i)
		print(col)
		print(sheet.cell(col,1).value)		
		if sheet.cell(col,1).value == currentDate:
			print(col)
			return col
			
			
Key = 'befefdd5573f442bb6374c64e17d1301'
CF.Key.set(Key)



connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()
print("1")
attend = [0 for i in range(60)]	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
	print("1")
	if filename.endswith(".jpg"):
		imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
		res = CF.face.detect(imgurl[3:])
		if len(res) != 1:
			print ("No face detected.")
			continue
			
		faceIds = []
		for face in res:
			faceIds.append(face['faceId'])
		res = CF.face.identify(faceIds, personGroupId)
		print (filename)
		print (res)
		for face  in res:
			if not face['candidates']:
				print ("Unknown")
			else:
				personId = face['candidates'][0]['personId']
				c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
				row = c.fetchone()
				attend[int(row[0])] += 1
				print (row[1] + " recognized")
		time.sleep(0.1)

for row in range(3, sheet.max_row+1):
	#rn = sheet.cell('A%s'% row).value
	rn = sheet.cell(row,column=1).value
	if rn is not None:
		rn = str(rn)[-2:]
		if attend[int(rn)] != 0:
			col = 'C' # getDateColumn()
			sheet['%s%s' % (col, str(row))] = 'P'
		else:
			col = 'C' # getDateColumn()
			sheet['%s%s' % (col, str(row))] = 'A'
wb.save(filename = 'reports%s.xlsx' % currentDate)	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
