
from openpyxl import load_workbook
from PDFWriter import PDFWriter
import cognitive_face as CF
personGroupId = 'test1'
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

import time
os.chdir("C:/Users/harsh/Desktop/DBA/project")

#get current date
currentDate = time.strftime("%d_%m_%y")
workbook = load_workbook(filename = 'reports%s.xlsx' % currentDate, guess_types=True, data_only=True)
sheet = wb.get_sheet_by_name('Cse15')

worksheet = workbook.active

pw = PDFWriter(filename = 'reports%s.pdf' % currentDate)
pw.setFont('Courier', 12)
pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
pw.setFooter('Generated using openpyxl and xtopdf')

ws_range = worksheet.iter_rows('A1:H13')
for row in ws_range:
    s = ''
    for cell in row:
        if cell.value is None:
            s += ' ' * 11
        else:
            s += str(cell.value).rjust(10) + ' '
    pw.writeLine(s)
pw.savePage()
pw.close()