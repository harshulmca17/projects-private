import cognitive_face as CF
personGroupId = 'test1'
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string

import time
os.chdir("C:/Users/harsh/Desktop/DBA/project")

#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = 'reports%s.xlsx' % currentDate)
sheet = wb.get_sheet_by_name('Cse15')

#print(type(currentDate))
			
Key = 'befefdd5573f442bb6374c64e17d1301'
CF.Key.set(Key)

absent = 0
present = 0
for row in range(3, sheet.max_row+1):
	#rn = sheet.cell('A%s'% row).value
	rn = sheet.cell(row,column=3).value
	if rn == 'A':
		absent+=1
	else:
		present+=1
wb.save(filename = 'reports%s.xlsx' % currentDate)
print(present)	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res
